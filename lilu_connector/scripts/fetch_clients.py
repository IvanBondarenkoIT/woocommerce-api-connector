"""
Скрипт для получения клиентов из LILU API и сохранения в файл.

Для Junior разработчиков:
Этот скрипт получает всех клиентов из системы LILU и сохраняет их
в файлы разных форматов (JSON, CSV, Excel).

Использование:
    python -m lilu_connector.scripts.fetch_clients
    python -m lilu_connector.scripts.fetch_clients --format json
    python -m lilu_connector.scripts.fetch_clients --format csv
    python -m lilu_connector.scripts.fetch_clients --format excel
"""

import sys
import os
import codecs
import json
import csv
from datetime import datetime
from typing import List, Dict, Any

# Fix encoding for Windows console
if sys.platform == 'win32':
    sys.stdout = codecs.getwriter('utf-8')(sys.stdout.buffer, 'strict')
    sys.stderr = codecs.getwriter('utf-8')(sys.stderr.buffer, 'strict')

# Добавляем корневую директорию проекта в путь для импорта
script_dir = os.path.dirname(os.path.abspath(__file__))
project_root = os.path.dirname(os.path.dirname(script_dir))
sys.path.insert(0, project_root)

# Загружаем .env файл из корня проекта
from dotenv import load_dotenv
env_path = os.path.join(project_root, '.env')
load_dotenv(env_path)

from lilu_connector import LILUConnector
from lilu_connector.models.client import ClientModel
from lilu_connector.api.exceptions import (
    AuthenticationError,
    NetworkError,
    ConfigurationError,
    LILUAPIError,
)
from lilu_connector.config.settings import LILUSettings


def save_to_json(clients: List[ClientModel], filename: str):
    """Сохранить клиентов в JSON файл."""
    data = [client.to_dict() for client in clients]
    
    with open(filename, 'w', encoding='utf-8') as f:
        json.dump(data, f, ensure_ascii=False, indent=2)
    
    print(f"✅ Сохранено {len(clients)} клиентов в {filename}")


def save_to_csv(clients: List[ClientModel], filename: str):
    """Сохранить клиентов в CSV файл."""
    if not clients:
        print("⚠️  Нет клиентов для сохранения")
        return
    
    fieldnames = list(clients[0].to_dict().keys())
    
    with open(filename, 'w', newline='', encoding='utf-8-sig') as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        
        for client in clients:
            writer.writerow(client.to_dict())
    
    print(f"✅ Сохранено {len(clients)} клиентов в {filename}")


def save_to_excel(clients: List[ClientModel], filename: str):
    """Сохранить клиентов в Excel файл."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        print("❌ ОШИБКА: Для сохранения в Excel требуется библиотека openpyxl")
        print("💡 Установите её командой: pip install openpyxl")
        return
    
    if not clients:
        print("⚠️  Нет клиентов для сохранения")
        return
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Клиенты"
    
    headers = list(clients[0].to_dict().keys())
    
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    for row, client in enumerate(clients, 2):
        client_dict = client.to_dict()
        for col, header in enumerate(headers, 1):
            value = client_dict.get(header, '')
            if isinstance(value, (dict, list)):
                value = json.dumps(value, ensure_ascii=False)
            ws.cell(row=row, column=col, value=value)
    
    for col in range(1, len(headers) + 1):
        column_letter = openpyxl.utils.get_column_letter(col)
        max_length = 0
        for row in ws[column_letter]:
            try:
                if len(str(row.value)) > max_length:
                    max_length = len(str(row.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    wb.save(filename)
    print(f"✅ Сохранено {len(clients)} клиентов в {filename}")


def fetch_and_save_clients(format_type: str = 'json'):
    """Получить клиентов из LILU API и сохранить в файл."""
    
    print("=" * 80)
    print("ПОЛУЧЕНИЕ И СОХРАНЕНИЕ КЛИЕНТОВ ИЗ LILU API")
    print("=" * 80)
    print()
    
    try:
        print("📋 Шаг 1: Инициализация коннектора...")
        try:
            connector = LILUConnector()
        except ValueError as e:
            raise ConfigurationError(str(e))
        
        print("✅ Коннектор инициализирован")
        print(f"   URL: {connector.settings.api_url}")
        print()
        
        print("📋 Шаг 2: Проверка подключения к API...")
        if connector.health_check():
            print("✅ API доступен")
        else:
            print("⚠️  API недоступен, но продолжаем...")
        print()
        
        print("📋 Шаг 3: Получение клиентов из API...")
        print("   Это может занять некоторое время...")
        print()
        
        all_clients = []
        page = 1
        limit = 50
        
        while True:
            try:
                print(f"   Получение страницы {page} (по {limit} клиентов)...", end=" ")
                clients = connector.get_clients(page=page, limit=limit)
                
                if not clients:
                    print("нет данных")
                    break
                
                all_clients.extend(clients)
                print(f"получено {len(clients)} клиентов")
                
                if len(clients) < limit:
                    break
                
                page += 1
                
            except Exception as e:
                print(f"ошибка: {e}")
                break
        
        print()
        print(f"✅ Всего получено клиентов: {len(all_clients)}")
        print()
        
        if not all_clients:
            print("⚠️  Клиенты не найдены. Нечего сохранять.")
            connector.close()
            return
        
        print("📊 Статистика:")
        active_count = sum(1 for c in all_clients if c.is_active)
        print(f"   Всего клиентов: {len(all_clients)}")
        print(f"   Активных: {active_count}")
        print(f"   Неактивных: {len(all_clients) - active_count}")
        print()
        
        print("📋 Шаг 4: Сохранение клиентов в файл...")
        
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_dir = os.path.join(project_root, "data", "output")
        os.makedirs(output_dir, exist_ok=True)
        
        if format_type.lower() == 'json':
            filename = os.path.join(output_dir, f"clients_{timestamp}.json")
            save_to_json(all_clients, filename)
        
        elif format_type.lower() == 'csv':
            filename = os.path.join(output_dir, f"clients_{timestamp}.csv")
            save_to_csv(all_clients, filename)
        
        elif format_type.lower() == 'excel':
            filename = os.path.join(output_dir, f"clients_{timestamp}.xlsx")
            save_to_excel(all_clients, filename)
        
        else:
            print(f"❌ Неизвестный формат: {format_type}")
            print("💡 Используйте: json, csv или excel")
            connector.close()
            return
        
        print()
        print("=" * 80)
        print("✅ ОПЕРАЦИЯ ЗАВЕРШЕНА УСПЕШНО")
        print("=" * 80)
        print()
        print(f"📁 Файл сохранён: {filename}")
        print(f"📊 Всего клиентов: {len(all_clients)}")
        print()
        
        if all_clients:
            print("📋 Первые 5 клиентов:")
            for i, client in enumerate(all_clients[:5], 1):
                print(f"   {i}. {client.name} ({client.email}) - {'Активен' if client.is_active else 'Неактивен'}")
        
        connector.close()
        
    except ConfigurationError as e:
        print("❌ ОШИБКА КОНФИГУРАЦИИ:")
        print(f"   {e}")
        print()
        print("💡 Решение:")
        print("   1. Убедитесь, что файл .env существует")
        print("   2. Проверьте, что в .env есть LILU_API_URL, LILU_API_KEY и LILU_API_SECRET")
        sys.exit(1)
    
    except AuthenticationError as e:
        print("❌ ОШИБКА АУТЕНТИФИКАЦИИ:")
        print(f"   {e}")
        print()
        print("💡 Решение:")
        print("   1. Проверьте правильность LILU_API_KEY и LILU_API_SECRET в .env")
        print("   2. Убедитесь, что ключи активны")
        sys.exit(1)
    
    except NetworkError as e:
        print("❌ ОШИБКА СЕТИ:")
        print(f"   {e}")
        print()
        print("💡 Решение:")
        print("   1. Проверьте интернет-соединение")
        print("   2. Проверьте правильность LILU_API_URL")
        sys.exit(1)
    
    except Exception as e:
        print("❌ НЕОЖИДАННАЯ ОШИБКА:")
        print(f"   {e}")
        print()
        import traceback
        traceback.print_exc()
        sys.exit(1)


if __name__ == "__main__":
    format_type = 'json'
    
    if len(sys.argv) > 1:
        if '--format' in sys.argv:
            idx = sys.argv.index('--format')
            if idx + 1 < len(sys.argv):
                format_type = sys.argv[idx + 1]
        elif sys.argv[1] in ['json', 'csv', 'excel']:
            format_type = sys.argv[1]
    
    fetch_and_save_clients(format_type)
