"""
WooCommerce GUI Application
Modern GUI for viewing and managing WooCommerce products
"""

import customtkinter as ctk
from tkinter import messagebox, ttk, filedialog
import json
import re
import threading
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from woocommerce_connector import WooCommerceConnector

# Set appearance mode and color theme
ctk.set_appearance_mode("dark")  # "light" or "dark"
ctk.set_default_color_theme("blue")  # "blue", "green", "dark-blue"


class WooCommerceGUI:
    """Main GUI application for WooCommerce"""
    
    def __init__(self):
        self.root = ctk.CTk()
        self.root.title("WooCommerce Product Manager")
        self.root.geometry("1200x800")
        
        # Initialize connector
        self.connector = None
        self.products = []
        self.current_product = None
        
        # Setup UI
        self.setup_ui()
        
        # Try to connect on startup
        self.connect_to_store()
    
    def setup_ui(self):
        """Setup the user interface"""
        # Main container
        main_frame = ctk.CTkFrame(self.root)
        main_frame.pack(fill="both", expand=True, padx=10, pady=10)
        
        # Top bar with connection status and refresh button
        top_bar = ctk.CTkFrame(main_frame)
        top_bar.pack(fill="x", padx=10, pady=10)
        
        self.status_label = ctk.CTkLabel(
            top_bar, 
            text="Status: Not Connected",
            font=ctk.CTkFont(size=14, weight="bold")
        )
        self.status_label.pack(side="left", padx=10)
        
        export_btn = ctk.CTkButton(
            top_bar,
            text="📊 Export to Excel",
            command=self.export_to_excel,
            width=140,
            fg_color="green",
            hover_color="darkgreen"
        )
        export_btn.pack(side="right", padx=5)
        
        refresh_btn = ctk.CTkButton(
            top_bar,
            text="🔄 Refresh",
            command=self.refresh_products,
            width=120
        )
        refresh_btn.pack(side="right", padx=5)
        
        # Main content area (split view)
        content_frame = ctk.CTkFrame(main_frame)
        content_frame.pack(fill="both", expand=True, padx=10, pady=10)
        
        # Left panel - Product list
        left_panel = ctk.CTkFrame(content_frame, width=400)
        left_panel.pack(side="left", fill="both", expand=False, padx=(0, 5))
        left_panel.pack_propagate(False)
        
        # Search bar
        search_frame = ctk.CTkFrame(left_panel)
        search_frame.pack(fill="x", padx=10, pady=10)
        
        self.search_entry = ctk.CTkEntry(
            search_frame,
            placeholder_text="Search products...",
            width=300
        )
        self.search_entry.pack(side="left", padx=5)
        self.search_entry.bind("<KeyRelease>", self.filter_products)
        
        # Product list with scrollbar
        list_frame = ctk.CTkFrame(left_panel)
        list_frame.pack(fill="both", expand=True, padx=10, pady=10)
        
        # Scrollable frame for products
        self.products_scroll = ctk.CTkScrollableFrame(list_frame)
        self.products_scroll.pack(fill="both", expand=True)
        
        # Right panel - Product details
        right_panel = ctk.CTkFrame(content_frame)
        right_panel.pack(side="right", fill="both", expand=True, padx=(5, 0))
        
        # Product details header
        details_header = ctk.CTkFrame(right_panel)
        details_header.pack(fill="x", padx=10, pady=10)
        
        self.product_title = ctk.CTkLabel(
            details_header,
            text="Select a product to view details",
            font=ctk.CTkFont(size=20, weight="bold")
        )
        self.product_title.pack(side="left", padx=10)
        
        # Scrollable details area
        details_scroll = ctk.CTkScrollableFrame(right_panel)
        details_scroll.pack(fill="both", expand=True, padx=10, pady=10)
        
        self.details_frame = ctk.CTkFrame(details_scroll)
        self.details_frame.pack(fill="both", expand=True)
        
        # Info label
        self.info_label = ctk.CTkLabel(
            self.details_frame,
            text="No product selected",
            font=ctk.CTkFont(size=14)
        )
        self.info_label.pack(pady=50)
    
    def connect_to_store(self):
        """Connect to WooCommerce store"""
        try:
            self.connector = WooCommerceConnector()
            self.status_label.configure(
                text="Status: ✅ Connected",
                text_color="green"
            )
            self.refresh_products()
        except Exception as e:
            self.status_label.configure(
                text=f"Status: ❌ Connection Failed",
                text_color="red"
            )
            messagebox.showerror("Connection Error", f"Failed to connect:\n{str(e)}")
    
    def refresh_products(self):
        """Refresh product list from store"""
        if not self.connector:
            messagebox.showwarning("Not Connected", "Please connect to store first")
            return
        
        # Show loading
        self.status_label.configure(text="Status: 🔄 Loading products...")
        
        # Load products in background thread
        def load_products():
            try:
                # Load all products with pagination
                self.root.after(0, lambda: self.status_label.configure(
                    text="Status: 🔄 Loading all products..."
                ))
                
                # Use get_all_products method which handles pagination automatically
                all_products = self.connector.get_all_products(per_page=100)
                
                if all_products:
                    self.products = all_products
                    # Update UI in main thread
                    self.root.after(0, self.update_product_list)
                else:
                    self.root.after(0, lambda: self.status_label.configure(
                        text="Status: ❌ No products found",
                        text_color="red"
                    ))
            except Exception as e:
                self.root.after(0, lambda: messagebox.showerror(
                    "Error",
                    f"Failed to load products:\n{str(e)}"
                ))
                self.root.after(0, lambda: self.status_label.configure(
                    text="Status: ❌ Error",
                    text_color="red"
                ))
        
        threading.Thread(target=load_products, daemon=True).start()
    
    def update_product_list(self):
        """Update the product list display"""
        # Clear existing products
        for widget in self.products_scroll.winfo_children():
            widget.destroy()
        
        if not self.products:
            no_products = ctk.CTkLabel(
                self.products_scroll,
                text="No products found",
                font=ctk.CTkFont(size=14)
            )
            no_products.pack(pady=20)
            return
        
        # Display products
        for product in self.products:
            self.create_product_card(product)
        
        self.status_label.configure(
            text=f"Status: ✅ Connected ({len(self.products)} products)",
            text_color="green"
        )
    
    def create_product_card(self, product):
        """Create a product card in the list"""
        card = ctk.CTkFrame(self.products_scroll)
        card.pack(fill="x", padx=5, pady=5)
        
        # Product name
        name_label = ctk.CTkLabel(
            card,
            text=product.get('name', 'Unnamed Product'),
            font=ctk.CTkFont(size=14, weight="bold"),
            anchor="w"
        )
        name_label.pack(fill="x", padx=10, pady=(10, 5))
        
        # Product info row
        info_frame = ctk.CTkFrame(card)
        info_frame.pack(fill="x", padx=10, pady=5)
        
        # Price
        price = product.get('price', '0')
        regular_price = product.get('regular_price', '0')
        on_sale = product.get('on_sale', False)
        
        if on_sale and regular_price:
            price_text = f"💰 {price} ₾ (was {regular_price} ₾)"
            price_color = "green"
        else:
            price_text = f"💰 {price} ₾"
            price_color = "white"
        
        price_label = ctk.CTkLabel(
            info_frame,
            text=price_text,
            font=ctk.CTkFont(size=12),
            text_color=price_color
        )
        price_label.pack(side="left", padx=5)
        
        # Stock status
        stock_status = product.get('stock_status', 'unknown')
        stock_color = "green" if stock_status == "instock" else "red"
        stock_label = ctk.CTkLabel(
            info_frame,
            text=f"📦 {stock_status.upper()}",
            font=ctk.CTkFont(size=11),
            text_color=stock_color
        )
        stock_label.pack(side="right", padx=5)
        
        # View button
        view_btn = ctk.CTkButton(
            card,
            text="View Details",
            command=lambda p=product: self.show_product_details(p),
            width=100,
            height=30
        )
        view_btn.pack(pady=10)
    
    def filter_products(self, event=None):
        """Filter products based on search"""
        search_term = self.search_entry.get().lower()
        
        # Clear and rebuild list with filtered products
        for widget in self.products_scroll.winfo_children():
            widget.destroy()
        
        if not search_term:
            # Show all products
            for product in self.products:
                self.create_product_card(product)
        else:
            # Filter products
            filtered = [
                p for p in self.products
                if search_term in p.get('name', '').lower() or
                   search_term in p.get('sku', '').lower()
            ]
            
            if not filtered:
                no_results = ctk.CTkLabel(
                    self.products_scroll,
                    text="No products found",
                    font=ctk.CTkFont(size=14)
                )
                no_results.pack(pady=20)
            else:
                for product in filtered:
                    self.create_product_card(product)
    
    def show_product_details(self, product):
        """Show detailed product information"""
        self.current_product = product
        
        # Clear details frame
        for widget in self.details_frame.winfo_children():
            widget.destroy()
        
        # Product title
        title = ctk.CTkLabel(
            self.details_frame,
            text=product.get('name', 'Unnamed Product'),
            font=ctk.CTkFont(size=24, weight="bold")
        )
        title.pack(pady=10)
        
        # Product ID
        id_label = ctk.CTkLabel(
            self.details_frame,
            text=f"ID: {product.get('id', 'N/A')}",
            font=ctk.CTkFont(size=12),
            text_color="gray"
        )
        id_label.pack(pady=5)
        
        # Separator
        separator = ctk.CTkFrame(self.details_frame, height=2, fg_color="gray")
        separator.pack(fill="x", pady=10)
        
        # Price information
        price_frame = ctk.CTkFrame(self.details_frame)
        price_frame.pack(fill="x", padx=20, pady=10)
        
        regular_price = product.get('regular_price', '0')
        sale_price = product.get('sale_price')
        on_sale = product.get('on_sale', False)
        
        if on_sale and sale_price:
            price_text = f"Sale Price: {sale_price} ₾"
            regular_text = f"Regular Price: {regular_price} ₾"
            price_color = "green"
        else:
            price_text = f"Price: {regular_price} ₾"
            regular_text = ""
            price_color = "white"
        
        price_label = ctk.CTkLabel(
            price_frame,
            text=price_text,
            font=ctk.CTkFont(size=18, weight="bold"),
            text_color=price_color
        )
        price_label.pack(side="left")
        
        if regular_text:
            regular_label = ctk.CTkLabel(
                price_frame,
                text=regular_text,
                font=ctk.CTkFont(size=14),
                text_color="gray"
            )
            regular_label.pack(side="right")
        
        # Stock information
        stock_frame = ctk.CTkFrame(self.details_frame)
        stock_frame.pack(fill="x", padx=20, pady=10)
        
        stock_status = product.get('stock_status', 'unknown')
        stock_quantity = product.get('stock_quantity')
        stock_color = "green" if stock_status == "instock" else "red"
        
        stock_text = f"Stock Status: {stock_status.upper()}"
        if stock_quantity is not None:
            stock_text += f" ({stock_quantity} units)"
        
        stock_label = ctk.CTkLabel(
            stock_frame,
            text=stock_text,
            font=ctk.CTkFont(size=14),
            text_color=stock_color
        )
        stock_label.pack(side="left")
        
        # Categories
        categories = product.get('categories', [])
        if categories:
            cat_frame = ctk.CTkFrame(self.details_frame)
            cat_frame.pack(fill="x", padx=20, pady=10)
            
            cat_label = ctk.CTkLabel(
                cat_frame,
                text="Categories:",
                font=ctk.CTkFont(size=14, weight="bold")
            )
            cat_label.pack(anchor="w", pady=5)
            
            cat_names = [cat.get('name', '') for cat in categories]
            cat_text = ", ".join(cat_names)
            cat_value = ctk.CTkLabel(
                cat_frame,
                text=cat_text,
                font=ctk.CTkFont(size=12),
                text_color="lightblue"
            )
            cat_value.pack(anchor="w")
        
        # Description
        description = product.get('description', '')
        short_desc = product.get('short_description', '')
        
        if description or short_desc:
            desc_frame = ctk.CTkFrame(self.details_frame)
            desc_frame.pack(fill="both", expand=True, padx=20, pady=10)
            
            desc_label = ctk.CTkLabel(
                desc_frame,
                text="Description:",
                font=ctk.CTkFont(size=14, weight="bold")
            )
            desc_label.pack(anchor="w", pady=5)
            
            desc_text = short_desc if short_desc else description
            if desc_text:
                # Remove HTML tags for display
                desc_text = re.sub('<[^<]+?>', '', desc_text)
                desc_text = desc_text[:500] + "..." if len(desc_text) > 500 else desc_text
                
                desc_value = ctk.CTkLabel(
                    desc_frame,
                    text=desc_text,
                    font=ctk.CTkFont(size=12),
                    justify="left",
                    wraplength=600
                )
                desc_value.pack(anchor="w", fill="x")
        
        # Product data (JSON view - collapsible)
        data_frame = ctk.CTkFrame(self.details_frame)
        data_frame.pack(fill="both", expand=True, padx=20, pady=10)
        
        data_label = ctk.CTkLabel(
            data_frame,
            text="Full Product Data:",
            font=ctk.CTkFont(size=14, weight="bold")
        )
        data_label.pack(anchor="w", pady=5)
        
        # JSON text area
        json_text = json.dumps(product, indent=2, ensure_ascii=False)
        json_textbox = ctk.CTkTextbox(data_frame, height=200)
        json_textbox.pack(fill="both", expand=True)
        json_textbox.insert("1.0", json_text)
        json_textbox.configure(state="disabled")
        
        # Future: Edit button (placeholder)
        edit_frame = ctk.CTkFrame(self.details_frame)
        edit_frame.pack(fill="x", padx=20, pady=20)
        
        edit_btn = ctk.CTkButton(
            edit_frame,
            text="✏️ Edit Product (Coming Soon)",
            command=self.edit_product_placeholder,
            width=200,
            height=40,
            state="disabled",
            fg_color="gray"
        )
        edit_btn.pack()
    
    def edit_product_placeholder(self):
        """Placeholder for future edit functionality"""
        messagebox.showinfo(
            "Coming Soon",
            "Product editing functionality will be available in the next version!"
        )
    
    def export_to_excel(self):
        """Export all products to Excel file with categories as sheets"""
        if not self.products:
            messagebox.showwarning("No Products", "Please load products first by clicking Refresh")
            return
        
        # Ask for file location
        filename = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
            initialfile=f"woocommerce_products_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        )
        
        if not filename:
            return  # User cancelled
        
        # Show progress in status
        self.status_label.configure(text="Status: 📊 Exporting to Excel...")
        
        # Export in background thread
        def export_thread():
            try:
                self._export_products_to_excel(filename)
                self.root.after(0, lambda: self.status_label.configure(
                    text=f"Status: ✅ Exported to Excel",
                    text_color="green"
                ))
                self.root.after(0, lambda: messagebox.showinfo(
                    "Export Complete",
                    f"Products exported successfully to:\n{filename}"
                ))
            except Exception as e:
                self.root.after(0, lambda: self.status_label.configure(
                    text="Status: ❌ Export Failed",
                    text_color="red"
                ))
                self.root.after(0, lambda: messagebox.showerror(
                    "Export Error",
                    f"Failed to export products:\n{str(e)}"
                ))
        
        threading.Thread(target=export_thread, daemon=True).start()
    
    def _export_products_to_excel(self, filename):
        """Internal method to export products to Excel"""
        wb = Workbook()
        wb.remove(wb.active)  # Remove default sheet
        
        # Get all unique categories
        categories_dict = {}
        products_without_category = []
        
        for product in self.products:
            categories = product.get('categories', [])
            if categories:
                # Add product to each category it belongs to
                for cat in categories:
                    cat_id = cat.get('id')
                    cat_name = cat.get('name', 'Uncategorized')
                    if cat_id not in categories_dict:
                        categories_dict[cat_id] = {
                            'name': cat_name,
                            'products': []
                        }
                    categories_dict[cat_id]['products'].append(product)
            else:
                products_without_category.append(product)
        
        # Create sheet for each category
        for cat_id, cat_data in categories_dict.items():
            sheet_name = self._sanitize_sheet_name(cat_data['name'])
            ws = wb.create_sheet(title=sheet_name)
            self._write_products_to_sheet(ws, cat_data['products'])
        
        # Create sheet for products without category
        if products_without_category:
            ws = wb.create_sheet(title="No Category")
            self._write_products_to_sheet(ws, products_without_category)
        
        # Save workbook
        wb.save(filename)
    
    def _sanitize_sheet_name(self, name):
        """Sanitize sheet name for Excel (max 31 chars, no special chars)"""
        # Excel sheet name limitations
        invalid_chars = ['\\', '/', '*', '?', ':', '[', ']']
        for char in invalid_chars:
            name = name.replace(char, '_')
        # Limit length
        return name[:31] if len(name) > 31 else name
    
    def _write_products_to_sheet(self, ws, products):
        """Write products to a worksheet with all attributes as columns"""
        if not products:
            return
        
        # Get all possible keys from all products
        all_keys = set()
        for product in products:
            all_keys.update(self._flatten_dict(product).keys())
        
        # Sort keys for consistent column order
        sorted_keys = sorted(all_keys)
        
        # Write header row
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        for col_idx, key in enumerate(sorted_keys, 1):
            cell = ws.cell(row=1, column=col_idx, value=key)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Write product data
        for row_idx, product in enumerate(products, 2):
            flat_product = self._flatten_dict(product)
            for col_idx, key in enumerate(sorted_keys, 1):
                value = flat_product.get(key, '')
                # Convert complex types to string
                if isinstance(value, (dict, list)):
                    value = json.dumps(value, ensure_ascii=False)
                elif value is None:
                    value = ''
                ws.cell(row=row_idx, column=col_idx, value=value)
        
        # Auto-adjust column widths
        for col_idx, key in enumerate(sorted_keys, 1):
            column_letter = get_column_letter(col_idx)
            max_length = 0
            # Check header
            max_length = max(max_length, len(str(key)))
            # Check data
            for row in ws[column_letter]:
                try:
                    if row.value:
                        max_length = max(max_length, len(str(row.value)))
                except:
                    pass
            # Set width with some padding
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Freeze header row
        ws.freeze_panes = 'A2'
    
    def _flatten_dict(self, d, parent_key='', sep='.'):
        """Flatten nested dictionary for Excel export"""
        items = []
        for k, v in d.items():
            new_key = f"{parent_key}{sep}{k}" if parent_key else k
            if isinstance(v, dict):
                items.extend(self._flatten_dict(v, new_key, sep=sep).items())
            elif isinstance(v, list):
                # Handle lists - convert to string or process items
                if v and isinstance(v[0], dict):
                    # List of dicts - convert to JSON string
                    items.append((new_key, json.dumps(v, ensure_ascii=False)))
                else:
                    # Simple list - join with comma
                    items.append((new_key, ', '.join(str(item) for item in v)))
            else:
                items.append((new_key, v))
        return dict(items)
    
    def run(self):
        """Start the GUI application"""
        self.root.mainloop()


def main():
    """Main entry point"""
    app = WooCommerceGUI()
    app.run()


if __name__ == "__main__":
    main()

