"""
Tests for Excel export functionality
"""

import pytest
import os
import json
from unittest.mock import Mock, patch, MagicMock
from openpyxl import load_workbook

# Mock customtkinter before importing woocommerce_gui
import sys
from unittest.mock import MagicMock as MockModule

# Create mock for customtkinter
mock_ctk = MockModule()
mock_ctk.CTk = MockModule
mock_ctk.CTkFrame = MockModule
mock_ctk.CTkLabel = MockModule
mock_ctk.CTkButton = MockModule
mock_ctk.CTkEntry = MockModule
mock_ctk.CTkScrollableFrame = MockModule
mock_ctk.CTkTextbox = MockModule
mock_ctk.CTkFont = MockModule
mock_ctk.set_appearance_mode = MockModule
mock_ctk.set_default_color_theme = MockModule

sys.modules['customtkinter'] = mock_ctk
sys.modules['tkinter.messagebox'] = MockModule()
sys.modules['tkinter.filedialog'] = MockModule()

from woocommerce_gui import WooCommerceGUI
from woocommerce_connector import WooCommerceConnector


class TestExcelExport:
    """Test cases for Excel export functionality"""
    
    @pytest.fixture
    def sample_products(self):
        """Sample product data for testing"""
        return [
            {
                'id': 1,
                'name': 'Test Product 1',
                'price': '10.00',
                'regular_price': '15.00',
                'sale_price': '10.00',
                'on_sale': True,
                'stock_status': 'instock',
                'stock_quantity': 5,
                'categories': [
                    {'id': 1, 'name': 'Category A', 'slug': 'category-a'}
                ],
                'sku': 'SKU001',
                'description': 'Test description',
                'short_description': 'Short desc'
            },
            {
                'id': 2,
                'name': 'Test Product 2',
                'price': '20.00',
                'regular_price': '20.00',
                'sale_price': '',
                'on_sale': False,
                'stock_status': 'outofstock',
                'stock_quantity': 0,
                'categories': [
                    {'id': 1, 'name': 'Category A', 'slug': 'category-a'},
                    {'id': 2, 'name': 'Category B', 'slug': 'category-b'}
                ],
                'sku': 'SKU002',
                'description': '',
                'short_description': ''
            },
            {
                'id': 3,
                'name': 'Product Without Category',
                'price': '30.00',
                'regular_price': '30.00',
                'categories': [],
                'sku': 'SKU003'
            }
        ]
    
    @pytest.fixture
    def mock_gui(self):
        """Create mock GUI instance"""
        # GUI is already mocked via sys.modules above
        gui = WooCommerceGUI()
        gui.products = []
        gui.status_label = MagicMock()
        return gui
    
    def test_flatten_dict_simple(self, mock_gui):
        """Test flattening simple dictionary"""
        data = {
            'id': 1,
            'name': 'Test',
            'price': '10.00'
        }
        
        result = mock_gui._flatten_dict(data)
        
        assert result['id'] == 1
        assert result['name'] == 'Test'
        assert result['price'] == '10.00'
    
    def test_flatten_dict_nested(self, mock_gui):
        """Test flattening nested dictionary"""
        data = {
            'id': 1,
            'dimensions': {
                'length': '10',
                'width': '20'
            }
        }
        
        result = mock_gui._flatten_dict(data)
        
        assert result['id'] == 1
        assert result['dimensions.length'] == '10'
        assert result['dimensions.width'] == '20'
    
    def test_flatten_dict_list(self, mock_gui):
        """Test flattening dictionary with list"""
        data = {
            'id': 1,
            'tags': ['tag1', 'tag2']
        }
        
        result = mock_gui._flatten_dict(data)
        
        assert result['id'] == 1
        assert 'tag1' in result['tags']
        assert 'tag2' in result['tags']
    
    def test_flatten_dict_list_of_dicts(self, mock_gui):
        """Test flattening dictionary with list of dicts"""
        data = {
            'id': 1,
            'categories': [
                {'id': 1, 'name': 'Cat1'},
                {'id': 2, 'name': 'Cat2'}
            ]
        }
        
        result = mock_gui._flatten_dict(data)
        
        assert result['id'] == 1
        assert isinstance(result['categories'], str)
        assert 'Cat1' in result['categories']
    
    def test_sanitize_sheet_name(self, mock_gui):
        """Test sheet name sanitization"""
        # Test long name
        long_name = 'A' * 50
        result = mock_gui._sanitize_sheet_name(long_name)
        assert len(result) <= 31
        
        # Test special characters
        special_name = 'Test/Sheet*Name?'
        result = mock_gui._sanitize_sheet_name(special_name)
        assert '/' not in result
        assert '*' not in result
        assert '?' not in result
    
    def test_export_products_to_excel(self, mock_gui, sample_products, tmp_path):
        """Test exporting products to Excel"""
        mock_gui.products = sample_products
        
        # Mock filedialog
        with patch('woocommerce_gui.filedialog.asksaveasfilename') as mock_dialog:
            filename = str(tmp_path / "test_export.xlsx")
            mock_dialog.return_value = filename
            
            # Mock status label update
            mock_gui.status_label = MagicMock()
            
            # Export
            mock_gui._export_products_to_excel(filename)
            
            # Verify file was created
            assert os.path.exists(filename)
            
            # Load and verify Excel file
            wb = load_workbook(filename)
            
            # Should have sheets for categories
            sheet_names = wb.sheetnames
            assert 'Category A' in sheet_names or 'No Category' in sheet_names
            
            # Verify data in first sheet
            ws = wb[sheet_names[0]]
            assert ws.max_row > 1  # Header + at least one product
    
    def test_export_grouping_by_category(self, mock_gui, sample_products, tmp_path):
        """Test that products are grouped by category in Excel"""
        mock_gui.products = sample_products
        
        filename = str(tmp_path / "test_grouping.xlsx")
        mock_gui._export_products_to_excel(filename)
        
        wb = load_workbook(filename)
        sheet_names = wb.sheetnames
        
        # Should have separate sheets for different categories
        # Category A should have products
        if 'Category A' in sheet_names:
            ws = wb['Category A']
            # Should have header + products
            assert ws.max_row >= 2
        
        # Products without category should be in separate sheet
        if 'No Category' in sheet_names:
            ws = wb['No Category']
            assert ws.max_row >= 2
    
    def test_export_all_attributes(self, mock_gui, sample_products, tmp_path):
        """Test that all product attributes are exported"""
        mock_gui.products = sample_products[:1]  # Just one product
        
        filename = str(tmp_path / "test_attributes.xlsx")
        mock_gui._export_products_to_excel(filename)
        
        wb = load_workbook(filename)
        ws = wb.active
        
        # Get header row
        headers = [cell.value for cell in ws[1]]
        
        # Should contain key attributes
        assert 'id' in headers
        assert 'name' in headers
        assert 'price' in headers
        assert 'stock_status' in headers
    
    def test_export_empty_products(self, mock_gui, tmp_path):
        """Test exporting when no products"""
        mock_gui.products = []
        
        filename = str(tmp_path / "test_empty.xlsx")
        
        # Should not raise error
        mock_gui._export_products_to_excel(filename)
        
        # File should still be created (with empty sheets)
        assert os.path.exists(filename)



